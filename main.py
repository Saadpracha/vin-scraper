import asyncio
import csv
import os
import random
import sys
from pathlib import Path

from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeout
)

# =========================
# CONFIG
# =========================

TIMEOUT_MS = 10000

# After each full batch, timed-out VINs run again in a fresh browser session.
MAX_TIMEOUT_RETRY_SESSIONS = 2

INPUT_FILE = "FORD-VIN-2020-2026.csv"

URL = "https://fr.ford.ca/support/recalls/"

OUTPUT_FILE = "fr_ford_recalls_output.csv"

# Runtime throttling + crash-resume
BATCH_SIZE = 50
PER_VIN_DELAY_SECONDS_RANGE = (2, 5)
BATCH_SLEEP_SECONDS_RANGE = (12 * 60, 15 * 60)
CHECKPOINT_FILE = "checkpoint_last_vin.txt"
FAILED_VINS_FILE = "failed_vins.csv"

# One random row per VIN; HTTP proxy (CONNECT) for Chromium.
# CSV columns: Host, Port, User, Pass (case-insensitive headers).
PROXIES_FILE = "iproyal-proxies-10.csv"

# Ford CA — contains(., ...) catches trailing * and mixed text nodes
XPATH_H2_NO_RECALLS = '//h2[contains(., "No Recalls")]'
XPATH_H2_NO_CSP = (
    '//h2[contains(., "No Customer Satisfaction Programs")]'
)
XPATH_H2_RECALLS_OR_CSP = (
    '//h2[contains(., "Recalls")] | '
    '//h2[contains(., "Customer Satisfaction Programs")]'
)

XPATH_VEHICLE_YM = '//div[@class="vehicle-information-ym"]'


# =========================
# HELPERS
# =========================

def _normalize_header(header):
    return (str(header or "")).strip().upper()


async def safe_text(locator):
    try:
        if await locator.count() > 0:
            return (await locator.first.inner_text()).strip()
    except Exception:
        pass
    return ""


def read_vins_from_file():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise ValueError(f"Input file not found: {INPUT_FILE}")

    suffix = input_path.suffix.lower()

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError(
                "Install openpyxl: pip install openpyxl"
            ) from exc

        workbook = load_workbook(
            filename=INPUT_FILE,
            read_only=True,
            data_only=True,
        )
        sheet = workbook.active
        header_row = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if not header_row:
            raise ValueError(f"No header row in {INPUT_FILE}")

        vin_col_index = None
        for index, header in enumerate(header_row):
            if _normalize_header(header) == "VIN":
                vin_col_index = index
                break
        if vin_col_index is None:
            raise ValueError(f"Column 'VIN' not found in {INPUT_FILE}")

        vins = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if vin_col_index >= len(row):
                continue
            vin = str(row[vin_col_index] or "").strip()
            if vin:
                vins.append(vin)
        workbook.close()
        if not vins:
            raise ValueError(
                f"No VIN values in column 'VIN' in {INPUT_FILE}"
            )
        print(f"Loaded {len(vins)} VIN(s) from xlsx: {INPUT_FILE}")
        return vins

    if suffix == ".csv":
        encodings_to_try = [
            "utf-8-sig", "utf-8", "cp1252", "latin-1"
        ]
        last_decode_error = None
        for encoding in encodings_to_try:
            try:
                with open(
                    INPUT_FILE, "r", newline="", encoding=encoding
                ) as csvfile:
                    reader = csv.DictReader(csvfile)
                    if not reader.fieldnames:
                        raise ValueError(
                            f"No header row in {INPUT_FILE}"
                        )
                    normalized = {
                        _normalize_header(h): h
                        for h in reader.fieldnames
                    }
                    vin_key = normalized.get("VIN")
                    if not vin_key:
                        raise ValueError(
                            f"Column 'VIN' not found in {INPUT_FILE}"
                        )
                    vins = [
                        (row.get(vin_key) or "").strip()
                        for row in reader
                    ]
                    vins = [v for v in vins if v]
                    if not vins:
                        raise ValueError(
                            f"No VIN values in column 'VIN' in {INPUT_FILE}"
                        )
                    print(
                        f"Loaded {len(vins)} VIN(s) from csv ({encoding})"
                    )
                    return vins
            except UnicodeDecodeError as decode_error:
                last_decode_error = decode_error
                continue
        if last_decode_error:
            raise ValueError(
                f"Could not decode {INPUT_FILE}."
            ) from last_decode_error
        raise ValueError(f"Could not read {INPUT_FILE}")

    raise ValueError(
        f"Unsupported extension '{suffix}'. Use .xlsx or .csv"
    )


def load_checkpoint_last_vin():
    """
    Returns the last successfully-scraped VIN (string) or "" if none.
    Kept as a single-line text file to avoid JSON corruption issues on crash.
    """
    path = Path(CHECKPOINT_FILE)
    if not path.exists():
        return ""
    try:
        return (path.read_text(encoding="utf-8") or "").strip()
    except Exception:
        return ""


def save_checkpoint_last_vin(vin):
    try:
        Path(CHECKPOINT_FILE).write_text(
            (vin or "").strip(),
            encoding="utf-8",
        )
    except Exception:
        pass


def log_failed_vin(vin, reason):
    """
    Append failures to a small CSV so you can re-run only failed VINs later.
    """
    file_exists = Path(FAILED_VINS_FILE).exists()
    try:
        with open(
            FAILED_VINS_FILE,
            "a",
            newline="",
            encoding="utf-8-sig",
        ) as f:
            writer = csv.DictWriter(
                f, fieldnames=["VIN", "Reason"]
            )
            if not file_exists:
                writer.writeheader()
            writer.writerow(
                {"VIN": (vin or "").strip(), "Reason": str(reason)}
            )
    except Exception:
        pass


def iter_vins_from_csv_streaming(resume_after_vin=""):
    """
    Stream VINs from INPUT_FILE without loading them all into RAM.

    If resume_after_vin is provided, skip rows until that VIN is seen once,
    then yield subsequent VINs.
    """
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise ValueError(f"Input file not found: {INPUT_FILE}")
    if input_path.suffix.lower() != ".csv":
        # For xlsx we keep the old behavior (small lists only).
        for vin in read_vins_from_file():
            yield vin
        return

    encodings_to_try = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_decode_error = None
    for encoding in encodings_to_try:
        try:
            with open(
                INPUT_FILE,
                "r",
                newline="",
                encoding=encoding,
            ) as csvfile:
                reader = csv.DictReader(csvfile)
                if not reader.fieldnames:
                    raise ValueError(f"No header row in {INPUT_FILE}")

                normalized = {
                    _normalize_header(h): h for h in reader.fieldnames
                }
                vin_key = normalized.get("VIN")
                if not vin_key:
                    raise ValueError(
                        f"Column 'VIN' not found in {INPUT_FILE}"
                    )

                skipping = bool(resume_after_vin)
                resume_after_vin = (resume_after_vin or "").strip()

                for row in reader:
                    vin = (row.get(vin_key) or "").strip()
                    if not vin:
                        continue
                    if skipping:
                        if vin == resume_after_vin:
                            skipping = False
                        continue
                    yield vin
            return
        except UnicodeDecodeError as decode_error:
            last_decode_error = decode_error
            continue
    if last_decode_error:
        raise ValueError(
            f"Could not decode {INPUT_FILE}."
        ) from last_decode_error
    raise ValueError(f"Could not read {INPUT_FILE}")


def _row_to_playwright_proxy(row):
    """
    Build Playwright `proxy` dict from a CSV row (normalized keys).
    Returns None if host/port missing.
    """
    norm = {
        _normalize_header(k): (v if v is not None else "")
        for k, v in row.items()
        if k is not None
    }
    host = str(norm.get("HOST", "")).strip()
    port = str(norm.get("PORT", "")).strip()
    user = str(norm.get("USER", "")).strip()
    password = str(norm.get("PASS", "")).strip()
    if not host or not port:
        return None
    proxy = {"server": f"http://{host}:{port}"}
    if user:
        proxy["username"] = user
    if password:
        proxy["password"] = password
    return proxy


def load_proxies_from_file():
    """
    Load proxy endpoints from PROXIES_FILE (CSV).
    Returns a list of Playwright-compatible proxy dicts; empty if
    file missing or no valid rows.
    """
    path = Path(PROXIES_FILE)
    if not path.exists():
        print(
            f"\nProxy file not found ({PROXIES_FILE}); "
            "scraping without proxy."
        )
        return []

    encodings_to_try = [
        "utf-8-sig", "utf-8", "cp1252", "latin-1"
    ]
    last_decode_error = None
    for encoding in encodings_to_try:
        try:
            with open(
                PROXIES_FILE,
                "r",
                newline="",
                encoding=encoding,
            ) as csvfile:
                reader = csv.DictReader(csvfile)
                if not reader.fieldnames:
                    print(
                        f"\nProxy file has no header row ({PROXIES_FILE}); "
                        "scraping without proxy."
                    )
                    return []
                proxies = []
                for row in reader:
                    p = _row_to_playwright_proxy(row)
                    if p:
                        proxies.append(p)
                if not proxies:
                    print(
                        f"\nNo valid proxy rows in {PROXIES_FILE}; "
                        "scraping without proxy."
                    )
                    return []
                print(
                    f"Loaded {len(proxies)} proxy endpoint(s) from "
                    f"{PROXIES_FILE}"
                )
                return proxies
        except UnicodeDecodeError as decode_error:
            last_decode_error = decode_error
            continue
    if last_decode_error:
        print(
            f"\nCould not decode {PROXIES_FILE}: {last_decode_error}; "
            "scraping without proxy."
        )
    return []


async def purge_cookies_and_cache(context, page):
    """
    Clear cookies and disk cache before closing the browser instance.
    """
    try:
        await context.clear_cookies()
    except Exception:
        pass
    try:
        cdp = await context.new_cdp_session(page)
        await cdp.send("Network.clearBrowserCache")
    except Exception:
        pass


def _token_looks_like_model_year(token):
    """True for a 4-digit calendar year typical on Ford VIN result pages."""
    if not token or len(token) != 4 or not token.isdigit():
        return False
    year = int(token)
    return 1980 <= year <= 2039


def split_year_and_model(vehicle_line):
    """
    Ford CA shows year + model in one string; order depends on locale.

    English: first token is the year (e.g. ``2024 Escape``).
    French: last token is the year; everything before is the model
    (e.g. ``Escape 2024``, ``Bronco Sport 2021``).
    """
    line = (vehicle_line or "").strip()
    if not line:
        return "", ""
    parts = line.split()
    if len(parts) == 1:
        only = parts[0]
        if _token_looks_like_model_year(only):
            return only, ""
        return "", only

    first, last = parts[0], parts[-1]
    first_is_year = _token_looks_like_model_year(first)
    last_is_year = _token_looks_like_model_year(last)

    if first_is_year:
        return first, " ".join(parts[1:])
    if last_is_year:
        return last, " ".join(parts[:-1])

    return "", line


async def expand_all_collapsed_accordions(page):
    """
    When recall or CSP has data, expand every collapsed row first.
    Pattern: accordion-item/button[@aria-expanded="false"] under #recalls-content.
    """
    root = page.locator("#recalls-content")
    if await root.count() == 0:
        root = page

    collapsed_xpath = (
        './/div[contains(@class,"accordion-item")]'
        '/button[@aria-expanded="false"]'
    )

    for _ in range(50):
        collapsed = root.locator(f"xpath={collapsed_xpath}")
        if await collapsed.count() == 0:
            break
        btn = collapsed.first
        try:
            await btn.scroll_into_view_if_needed()
            await btn.click(timeout=TIMEOUT_MS)
        except Exception:
            break


async def wait_for_results_panel_ready(page):
    """Any recalls/CSP shell or heading suffices; attached (not visible)."""
    try:
        await page.locator("#recalls-content").wait_for(
            state="attached",
            timeout=TIMEOUT_MS,
        )
    except PlaywrightTimeout:
        pass

    ready_marker = (
        page.locator("#recalls-section")
        .or_(page.locator("#csp-section"))
        .or_(page.locator("#recalls-content"))
        .or_(
            page.locator("#recalls-content div.recalls-section").first
        )
        .or_(page.locator("#recalls-content div.csp-section").first)
        .or_(page.locator(f"xpath={XPATH_H2_RECALLS_OR_CSP}"))
        .or_(page.locator(f"xpath={XPATH_H2_NO_RECALLS}"))
        .or_(page.locator(f"xpath={XPATH_H2_NO_CSP}"))
    )
    await ready_marker.first.wait_for(
        state="attached",
        timeout=TIMEOUT_MS,
    )


async def locate_recalls_root(page):
    loc = page.locator("#recalls-section")
    if await loc.count() > 0:
        return loc
    accordions = page.locator(
        '#recalls-content [data-testid="recalls-csp-accordions"]'
    )
    nested = accordions.locator("div.recalls-section")
    if await nested.count() > 0:
        return nested.first
    loose = page.locator("#recalls-content div.recalls-section")
    if await loose.count() > 0:
        return loose.first
    return loc


async def locate_csp_root(page):
    loc = page.locator("#csp-section")
    if await loc.count() > 0:
        return loc
    accordions = page.locator(
        '#recalls-content [data-testid="recalls-csp-accordions"]'
    )
    nested = accordions.locator("div.csp-section")
    if await nested.count() > 0:
        return nested.first
    loose = page.locator("#recalls-content div.csp-section")
    if await loose.count() > 0:
        return loose.first
    return loc


async def load_recall_and_csp_data(page):
    await wait_for_results_panel_ready(page)

    recalls_shell = await locate_recalls_root(page)
    csp_shell = await locate_csp_root(page)

    has_recalls_shell = (await recalls_shell.count()) > 0
    has_csp_shell = (await csp_shell.count()) > 0

    no_recalls_heading = (
        await page.locator(f"xpath={XPATH_H2_NO_RECALLS}").count()
    ) > 0
    no_csp_heading = (
        await page.locator(f"xpath={XPATH_H2_NO_CSP}").count()
    ) > 0

    no_recalls_class = False
    if has_recalls_shell:
        root_cls = (
            await recalls_shell.first.get_attribute("class") or ""
        )
        no_recalls_class = (
            "no-recalls" in root_cls
            or (await recalls_shell.locator(".no-recalls").count()) > 0
        )

    no_csp_class = False
    if has_csp_shell:
        root_cls = await csp_shell.first.get_attribute("class") or ""
        no_csp_class = (
            "no-csp" in root_cls
            or (await csp_shell.locator(".no-csp").count()) > 0
        )

    no_recalls_state = no_recalls_heading or no_recalls_class
    no_csp_state = no_csp_heading or no_csp_class

    recall_or_csp_has_data = (
        (has_recalls_shell and not no_recalls_state)
        or (
            has_csp_shell and not no_csp_state
        )
    )
    if recall_or_csp_has_data:
        await expand_all_collapsed_accordions(page)

    if has_recalls_shell and not no_recalls_state:
        recalls = await parse_recalls(page)
    else:
        recalls = []

    if has_csp_shell and not no_csp_state:
        customer_satisfaction_programs = await parse_csp(page)
    else:
        customer_satisfaction_programs = []

    return recalls, customer_satisfaction_programs


# =========================
# PARSE RECALLS
# =========================

async def parse_recalls(page):
    """
    XPath semantics (Ford CA):
      title: #recalls-section//div[@class="accordion-titles"]/div
      status: #recalls-section//div[@class="accordion-titles"]/span
      campaign: div[@data-testid="accordion-description"] @id
    Description duplicates title.
    """
    recalls = []

    recalls_section = page.locator('#recalls-section')
    if await recalls_section.count() == 0:
        recalls_section = await locate_recalls_root(page)

    if await recalls_section.count() == 0:
        return recalls

    items = recalls_section.locator(
        'xpath=.//div[contains(@class,"accordion-item")]'
    )
    n = await items.count()

    for i in range(n):
        item = items.nth(i)

        title = await safe_text(
            item.locator(
                'xpath=.//div[@class="accordion-titles"]/div'
            )
        )
        title = " ".join(title.split()) if title else ""

        if not title:
            continue

        status = await safe_text(
            item.locator(
                'xpath=.//div[@class="accordion-titles"]/span'
            )
        )
        status = " ".join(status.split()) if status else ""

        campaign_el = item.locator(
            'xpath=.//div[@data-testid="accordion-description"]'
        )
        campaign = ""
        if await campaign_el.count() > 0:
            campaign = (
                await campaign_el.first.get_attribute("id") or ""
            ).strip()

        recalls.append({
            "title": title,
            "description": title,
            "campaign": campaign,
            "status": status
        })

    return recalls


# =========================
# PARSE CSP
# =========================

async def parse_csp(page):
    """
    XPath semantics:
      title: #csp-section//div[@class="accordion-titles"]/div
      campaign: div[data-testid=accordion-description] @id (per row)
      next step: div[contains(., Next Steps)]/following-sibling::div (per row)
    Description duplicates title (no CSP status).
    """
    csp_results = []

    csp_section = page.locator('#csp-section')
    if await csp_section.count() == 0:
        csp_section = await locate_csp_root(page)

    if await csp_section.count() == 0:
        return csp_results

    items = csp_section.locator(
        'xpath=.//div[contains(@class,"accordion-item")]'
    )
    n = await items.count()

    for i in range(n):
        item = items.nth(i)

        title = await safe_text(
            item.locator(
                'xpath=.//div[@class="accordion-titles"]/div'
            )
        )
        title = " ".join(title.split()) if title else ""

        if not title:
            continue

        campaign_el = item.locator(
            'xpath=.//div[@data-testid="accordion-description"]'
        )
        campaign = ""
        if await campaign_el.count() > 0:
            campaign = (
                await campaign_el.first.get_attribute("id") or ""
            ).strip()

        next_steps = await safe_text(
            item.locator(
                'xpath=.//div[contains(.,"Next Steps")]'
                '/following-sibling::div'
            )
        )
        next_steps = " ".join(next_steps.split()) if next_steps else ""

        csp_results.append({
            "title": title,
            "description": title,
            "campaign": campaign,
            "next_steps": next_steps
        })

    return csp_results


# =========================
# SAVE CSV
# =========================

def save_to_csv(data):
    file_exists = Path(OUTPUT_FILE).exists()
    rows = []

    if data["recalls"]:
        for recall in data["recalls"]:
            rows.append({
                "VIN": data["vin"],
                "Section": "Recall",
                "Year": data["year"],
                "Model": data["model"],
                "Title": recall["title"],
                "Description": recall["description"],
                "Campaign": recall["campaign"],
                "Status": recall["status"],
                "Next Steps": ""
            })
    else:
        rows.append({
            "VIN": data["vin"],
            "Section": "Recall",
            "Year": data["year"],
            "Model": data["model"],
            "Title": "",
            "Description": "",
            "Campaign": "",
            "Status": "No Recalls",
            "Next Steps": ""
        })

    if data["customer_satisfaction_programs"]:
        for csp in data["customer_satisfaction_programs"]:
            rows.append({
                "VIN": data["vin"],
                "Section": "Customer Satisfaction Program",
                "Year": data["year"],
                "Model": data["model"],
                "Title": csp["title"],
                "Description": csp["description"],
                "Campaign": csp["campaign"],
                "Status": "",
                "Next Steps": csp["next_steps"]
            })
    else:
        rows.append({
            "VIN": data["vin"],
            "Section": "Customer Satisfaction Program",
            "Year": data["year"],
            "Model": data["model"],
            "Title": "",
            "Description": "",
            "Campaign": "",
            "Status": "",
            "Next Steps": "No CSP"
        })

    fieldnames = [
        "VIN",
        "Section",
        "Year",
        "Model",
        "Title",
        "Description",
        "Campaign",
        "Status",
        "Next Steps",
    ]

    with open(
        OUTPUT_FILE,
        "a",
        newline="",
        encoding="utf-8-sig",
    ) as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)

    print(f"\nSaved data to: {OUTPUT_FILE}")


# =========================
# MAIN
# =========================

async def process_one_vin_fresh_browser(playwright, vin, proxy=None):
    """
    Dedicated browser lifecycle per VIN: launch → extract → purge
    cookies/cache → close (then next VIN starts a brand-new browser).

    `proxy`: optional Playwright proxy dict (e.g. random pick per VIN).

    Returns:
      - "ok" on success (saved to CSV + checkpoint updated)
      - "timeout" on PlaywrightTimeout (not saved)
      - "error" on other exceptions (not saved)
    """
    browser = await playwright.chromium.launch(
        headless=False,
        slow_mo=50,
    )

    context_options = {}
    if proxy:
        context_options["proxy"] = proxy
        print(
            "Browser context proxy: "
            f"{proxy.get('server', '')}"
        )

    context = await browser.new_context(**context_options)
    page = await context.new_page()

    status = "error"

    try:
        print(f"\nOpening: {URL}")

        await page.goto(URL, wait_until="domcontentloaded")

        vin_input = page.locator(
            '//input[@data-testid="input-text"]'
        )
        await vin_input.wait_for(
            state="visible",
            timeout=TIMEOUT_MS,
        )

        print(f"VIN lookup: {vin}")

        await vin_input.click()
        await vin_input.fill(vin)
        await vin_input.press("Enter")

        try:
            await page.wait_for_load_state(
                "networkidle",
                timeout=TIMEOUT_MS,
            )
        except PlaywrightTimeout:
            pass

        vehicle_info = page.locator(
            f"xpath={XPATH_VEHICLE_YM}"
        )
        await vehicle_info.wait_for(
            state="visible",
            timeout=TIMEOUT_MS,
        )

        vehicle_text = (
            await vehicle_info.first.inner_text()
        ).strip()
        year_part, model_part = split_year_and_model(
            vehicle_text
        )

        final_data = {
            "year": year_part,
            "model": model_part,
            "vin": vin,
            "recalls": [],
            "customer_satisfaction_programs": [],
        }

        (
            final_data["recalls"],
            final_data[
                "customer_satisfaction_programs"
            ],
        ) = await load_recall_and_csp_data(page)

        save_to_csv(final_data)
        save_checkpoint_last_vin(vin)
        status = "ok"

    except PlaywrightTimeout:
        print(
            f"\nTimeout for VIN {vin} — not saved; "
            "will retry."
        )
        status = "timeout"

    except Exception as vin_error:
        print(
            f"\nFailed for VIN {vin} (not a timeout): "
            f"{vin_error}"
        )
        status = "error"

    finally:
        try:
            await purge_cookies_and_cache(context, page)
        except Exception:
            pass

        await page.close()
        await context.close()
        await browser.close()

        print(
            f"Closed browser for VIN {vin} "
            "(cookies/cache cleared)."
        )

    return status


async def main():
    proxies = load_proxies_from_file()
    last_vin = load_checkpoint_last_vin()
    if last_vin:
        print(
            f"\nResuming after last saved VIN: {last_vin}\n"
            f"(If that VIN is not found in the input file, scraping will "
            f"start from the beginning.)"
        )

    vins_iter = iter_vins_from_csv_streaming(
        resume_after_vin=last_vin
    )

    async with async_playwright() as p:
        processed_in_batch = 0
        total_processed = 0

        for vin in vins_iter:
            vin = (vin or "").strip()
            if not vin:
                continue

            total_processed += 1
            processed_in_batch += 1

            # Small delay between VINs
            await asyncio.sleep(
                random.uniform(*PER_VIN_DELAY_SECONDS_RANGE)
            )

            proxy = random.choice(proxies) if proxies else None

            # Retry loop per VIN (fresh browser each attempt)
            last_error = ""
            status = "error"
            for attempt in range(1, MAX_TIMEOUT_RETRY_SESSIONS + 2):
                print(
                    f"\nVIN {vin} — attempt {attempt}/"
                    f"{MAX_TIMEOUT_RETRY_SESSIONS + 1}"
                )
                status = await process_one_vin_fresh_browser(
                    p, vin, proxy
                )
                if status == "ok":
                    break
                if status == "timeout":
                    last_error = "timeout"
                    await asyncio.sleep(random.uniform(5, 10))
                    continue
                last_error = "error"
                await asyncio.sleep(random.uniform(10, 20))

            if status != "ok":
                log_failed_vin(vin, last_error or status)

            # Batch sleep after every BATCH_SIZE VINs attempted
            if processed_in_batch >= BATCH_SIZE:
                processed_in_batch = 0
                sleep_s = random.uniform(*BATCH_SLEEP_SECONDS_RANGE)
                print(
                    f"\nBatch complete ({BATCH_SIZE} VINs). Sleeping "
                    f"{int(sleep_s)} seconds before continuing..."
                )
                await asyncio.sleep(sleep_s)

    print("\nDone.")


if __name__ == "__main__":
    _virtual_display = None
    if sys.platform.startswith("linux") and not os.environ.get("DISPLAY"):
        from pyvirtualdisplay import Display

        _virtual_display = Display(visible=0, size=(1920, 1080))
        _virtual_display.start()

    try:
        asyncio.run(main())
    finally:
        if _virtual_display is not None:
            _virtual_display.stop()
